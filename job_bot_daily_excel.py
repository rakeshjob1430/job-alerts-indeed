import os
import re
import time
import ssl
import smtplib
from datetime import datetime
from typing import List, Dict, Any

import requests
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font


# ================== ENV (GitHub Secrets) ==================
SERPAPI_KEY = os.getenv("SERPAPI_KEY")

EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")  # Gmail App Password
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")
# ==========================================================

LOCATION = "United States"

# You said you mainly want Food Safety Supervisor roles
ROLE_KEYWORDS = [
    "Food Safety Supervisor",
    "Food Safety Supervisor FSQA",  # helps some searches
]

FOOD_HINTS = [
    "food", "foods", "food manufacturing", "food processing", "plant", "production",
    "warehouse", "haccp", "sqf", "fsqa", "gmp", "sanitation", "usda", "fda",
    "meat", "poultry", "dairy", "bakery", "beverage", "produce"
]

# Only these sources
ALLOWED_SOURCES = {"indeed", "linkedin", "glassdoor", "ziprecruiter"}


def validate_env():
    if not SERPAPI_KEY:
        raise ValueError("SERPAPI_KEY missing (GitHub Secret).")
    if not EMAIL_SENDER or not EMAIL_PASSWORD or not EMAIL_RECEIVER:
        raise ValueError("EMAIL_SENDER / EMAIL_PASSWORD / EMAIL_RECEIVER missing (GitHub Secrets).")


# ---------------- SerpAPI calls (retry/backoff) ----------------
def serpapi_google_jobs(query: str, location: str, num: int = 50) -> List[Dict[str, Any]]:
    params = {
        "engine": "google_jobs",
        "q": query,
        "location": location,
        "api_key": SERPAPI_KEY,
        "num": num,
    }

    retry_statuses = {429, 502, 503, 504}
    max_attempts = 5

    for attempt in range(1, max_attempts + 1):
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=30)

            if r.status_code in retry_statuses:
                time.sleep(2 ** attempt)
                continue

            r.raise_for_status()
            data = r.json()
            return data.get("jobs_results", []) or []

        except requests.RequestException:
            time.sleep(2 ** attempt)

    return []


def serpapi_google_jobs_listing(job_id: str) -> Dict[str, Any]:
    if not job_id:
        return {}

    params = {"engine": "google_jobs_listing", "job_id": job_id, "api_key": SERPAPI_KEY}
    retry_statuses = {429, 502, 503, 504}
    max_attempts = 4

    for attempt in range(1, max_attempts + 1):
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=30)

            if r.status_code in retry_statuses:
                time.sleep(2 ** attempt)
                continue

            if r.status_code != 200:
                return {}

            return r.json() or {}

        except requests.RequestException:
            time.sleep(2 ** attempt)

    return {}


# ---------------- Helpers ----------------
def normalize_source(via_value: str) -> str:
    """
    SerpAPI google_jobs returns "via": "Indeed", "LinkedIn", "ZipRecruiter", etc.
    We'll normalize and filter to only 4 sources.
    """
    s = (via_value or "").strip().lower()
    if "indeed" in s:
        return "Indeed"
    if "linkedin" in s:
        return "LinkedIn"
    if "glassdoor" in s:
        return "Glassdoor"
    if "ziprecruiter" in s:
        return "ZipRecruiter"
    return ""


def is_allowed_source(via_value: str) -> bool:
    s = (via_value or "").lower()
    return any(x in s for x in ALLOWED_SOURCES)


def safe_apply_link(job: Dict[str, Any]) -> str:
    links = job.get("related_links") or []
    if isinstance(links, list) and links:
        link = links[0].get("link") or ""
        return link if link.startswith("http") else "N/A"
    return "N/A"


def safe_apply_link_from_details(details: Dict[str, Any]) -> str:
    apply_options = details.get("apply_options") or []
    if isinstance(apply_options, list) and apply_options:
        link = apply_options[0].get("link") or ""
        return link if link.startswith("http") else "N/A"
    return "N/A"


def safe_pay(job: Dict[str, Any]) -> str:
    de = job.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("salary"):
        return str(de["salary"])

    ext = job.get("extensions") or []
    if isinstance(ext, list):
        for item in ext:
            if isinstance(item, str) and ("$" in item or "hour" in item.lower() or "year" in item.lower()):
                return item
    return "N/A"


def safe_pay_from_details(details: Dict[str, Any]) -> str:
    de = details.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("salary"):
        return str(de["salary"])
    return "N/A"


def safe_time_posted(job: Dict[str, Any]) -> str:
    de = job.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("posted_at"):
        return str(de["posted_at"])

    ext = job.get("extensions") or []
    if isinstance(ext, list):
        for item in ext:
            if isinstance(item, str) and (
                "ago" in item.lower() or "today" in item.lower() or
                "yesterday" in item.lower() or "posted" in item.lower()
            ):
                return item
    return "N/A"


def safe_time_posted_from_details(details: Dict[str, Any]) -> str:
    de = details.get("detected_extensions") or {}
    if isinstance(de, dict) and de.get("posted_at"):
        return str(de["posted_at"])
    return "N/A"


def posted_days(time_posted: str) -> int:
    if not time_posted or time_posted == "N/A":
        return 999

    s = time_posted.strip().lower()
    if "just posted" in s or "today" in s:
        return 0
    if "yesterday" in s:
        return 1

    m = re.search(r"(\d+)\s+hour", s)
    if m:
        return 0

    m = re.search(r"(\d+)\s+day", s)
    if m:
        return int(m.group(1))

    m = re.search(r"(\d+)\s+week", s)
    if m:
        return int(m.group(1)) * 7

    return 999


def looks_food_industry(job: Dict[str, Any]) -> bool:
    text = " ".join([
        str(job.get("title") or ""),
        str(job.get("company_name") or ""),
        str(job.get("description") or ""),
    ]).lower()
    return any(h in text for h in FOOD_HINTS)


def normalize_row(job: Dict[str, Any]) -> Dict[str, str]:
    job_id = job.get("job_id") or "N/A"

    title = job.get("title") or "N/A"
    company = job.get("company_name") or "N/A"
    location = job.get("location") or "N/A"

    source_norm = normalize_source(job.get("via") or "")
    pay = safe_pay(job)
    time_posted = safe_time_posted(job)
    apply_link = safe_apply_link(job)

    # Try details if any key field missing
    if job_id != "N/A" and (pay == "N/A" or time_posted == "N/A" or apply_link == "N/A"):
        details = serpapi_google_jobs_listing(job_id)
        if details:
            if pay == "N/A":
                pay = safe_pay_from_details(details) or pay
            if time_posted == "N/A":
                time_posted = safe_time_posted_from_details(details) or time_posted
            if apply_link == "N/A":
                apply_link = safe_apply_link_from_details(details) or apply_link

    return {
        "job_id": job_id,  # dedupe only
        "title": title,
        "company name": company,     # EXACT column name requested
        "pay": pay if pay else "N/A",
        "time posted": time_posted if time_posted else "N/A",
        "location": location,
        "source": source_norm if source_norm else "N/A",
        "link to apply": apply_link if apply_link else "N/A",
    }


def build_queries() -> List[str]:
    """
    Focus ONLY on Food Safety Supervisor.
    (You can add more query variations if you want more results.)
    """
    queries = []
    for role in ROLE_KEYWORDS:
        queries.append(f'"{role}" food')
        queries.append(f'"{role}" HACCP')
        queries.append(f'"{role}" SQF')
        queries.append(f'"{role}" FSQA')
        queries.append(f'"{role}" GMP')
        queries.append(f'"{role}" food manufacturing')
        queries.append(f'"{role}" food processing')
    return queries


def dedupe(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    out = []
    for r in rows:
        key = r.get("job_id") or (r.get("title","") + "|" + r.get("company name","") + "|" + r.get("location",""))
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def create_excel(rows: List[Dict[str, str]], filename: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # EXACT order requested
    headers = ["title", "company name", "pay", "time posted", "location", "source", "link to apply"]
    ws.append(headers)

    # Write rows
    for r in rows:
        ws.append([r.get(h, "N/A") for h in headers])

    # Make "link to apply" clickable
    link_col = headers.index("link to apply") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col)
        val = str(cell.value or "")
        if val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(color="0000FF", underline="single")

    wb.save(filename)
    return filename


def send_email_with_attachment(subject: str, body: str, attachment_path: str):
    msg = EmailMessage()
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_RECEIVER
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(attachment_path),
    )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)


def main():
    validate_env()

    all_rows: List[Dict[str, str]] = []

    for q in build_queries():
        jobs = serpapi_google_jobs(q, LOCATION, num=50)

        for job in jobs:
            # Must be from only 4 sources
            if not is_allowed_source(job.get("via") or ""):
                continue

            # Must look like food industry
            if not looks_food_industry(job):
                continue

            all_rows.append(normalize_row(job))

    all_rows = dedupe(all_rows)

    # Keep last 7 days
    all_rows = [r for r in all_rows if posted_days(r.get("time posted", "N/A")) <= 7]

    # Sort newest first (0 days first)
    all_rows.sort(key=lambda r: posted_days(r.get("time posted", "N/A")))

    today = datetime.now().strftime("%Y-%m-%d")
    excel_file = f"ajay_jobs_{today}.xlsx"  # name can be anything

    create_excel(all_rows, excel_file)

    subject = f"Daily Food Safety Supervisor Jobs - {today}"
    body = f"""Hi,

Attached is your daily Excel list for Food Safety Supervisor roles (last 7 days).
Sources included ONLY: Indeed, LinkedIn, Glassdoor, ZipRecruiter
Total jobs found: {len(all_rows)}

Excel columns order:
title, company name, pay, time posted, location, source, link to apply

Regards,
Job Bot
"""

    send_email_with_attachment(subject, body, excel_file)


if __name__ == "__main__":
    main()
